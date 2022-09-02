from bs4 import BeautifulSoup as bs
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import Chrome
from time import sleep
import pandas as pd
import glob
import logging
import requests as rq
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet._write_only import WriteOnlyCell
from openpyxl.styles import Font, Alignment
import urllib3
from datetime import datetime

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

green = '0000FF00'
yellow = '00FFFF00'
blue = '0000FFFF'

url2 = 'https://requerimento.inss.gov.br/'
url = 'https://requerimento.inss.gov.br/saginternet/pages/agendamento/selecionarServico.xhtml'
login = ''
senha = ''
serviço = ''

def main():
    driver = Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(url)

    wa = WebDriverWait(driver, 3600)

    wa.until(EC.presence_of_element_located((By.ID, 'username')))
    log = driver.find_element(By.ID, 'username')
    log.send_keys(login)

    wa.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:servicoDrop_input')))

    while True:
        try:
            wa = WebDriverWait(driver, 2)
            wa.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:servicoDrop_input')))

            serv = driver.find_element(By.ID, 'formAgendarConsultar:servicoDrop_input')
            serv.send_keys(serviço)
            sleep(1)
            serv.send_keys(Keys.RETURN)

            sleep(2)
            element = WebDriverWait(driver, 10)
            element.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')))

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(1)

            element.until(EC.element_to_be_clickable((By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')))
            av = driver.find_element(By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')
            av.click()
            break
        except Exception as e:
            driver.refresh()
            logging.error(e)

    element.until(EC.presence_of_element_located((By.ID, 'formSugestaoDesistenciaRequerimento:btnAvancarSugestaoDesistenciaRequerimento')))

    r = WebDriverWait(driver, 10)
    r.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:cpfInput')))

    cookies = driver.get_cookies()[0]
    sessionid = cookies['name']+'='+cookies['value']

    html = driver.page_source
    soup = bs(html, 'html.parser')

    TOKEN = soup.find('input', {'name':'DTPINFRA_TOKEN'})['value']
    JSF = soup.find('input', {'name':'javax.faces.ViewState'})['value']

    url_drive = driver.current_url
    cid = url_drive[-1]

    driver.quit()

    url_p = f'https://requerimento.inss.gov.br/saginternet/pages/agendamento/registrar/dadosRequerenteEntidadeConveniada.xhtml?cid=1'



    header = {
        'Accept': 'application/xml, text/xml, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Cookie': sessionid,
        'Faces-Request': 'partial/ajax',
        'Host': 'requerimento.inss.gov.br',
        'Origin': 'https://requerimento.inss.gov.br',
        'Referer': 'https://requerimento.inss.gov.br/saginternet/pages/agendamento/selecionarServico.xhtml',
        'sec-ch-ua': '".Not/A)Brand";v="99", "Google Chrome";v="103", "Chromium";v="103"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'
        }

    chaves = []

    arquivos_xlsx = glob.glob('read/*')
    last = arquivos_xlsx[-1]

    with open('path/sheet.txt', 'r+', encoding='UTF-8') as t:
        pla = int(t.read())

    for x in range(pla, len(arquivos_xlsx)):
        co = 1
        maior = []

        while True:
            try:
                planilha = pd.read_excel(arquivos_xlsx[x], dtype=str)
                cpf1 = planilha['CPF'].tolist()
                nasc1 = planilha['Nascimento'].tolist()
                idade1 = planilha['Idade'].tolist()
                maes = planilha['NOME_MAE'].tolist()
                nb1 = planilha['NB'].tolist()
                data_inde1 = planilha['DATA DO INDEFERIMENTO'].tolist()
                especies = planilha['ESPÉCIE'].tolist()
                motivos = planilha['MOTIVO DO INDEFERIMENTO'].tolist()
                ddd_cel = planilha['DDD_CEL0'].tolist()
                ddd_cel1 = planilha['DDD_CEL1'].tolist()
                ddd_cel2 = planilha['DDD_CEL2'].tolist()
                celular0 = planilha['CELULAR'].tolist()
                celular1 = planilha['CELULAR1'].tolist()
                celular2 = planilha['CELULAR2'].tolist()
                ddds = planilha['DDD0'].tolist()
                ddds1 = planilha['DDD1'].tolist()
                ddds2 = planilha['DDD2'].tolist()
                tel0 = planilha['TELEFONE0'].tolist()
                tel1 = planilha['TELEFONE1'].tolist()
                tel2 = planilha['TELEFONE2'].tolist()
                logra1 = planilha['EnderecoLogradouro'].tolist()
                logra2 = planilha['EnderecoLogradouro_ARQUIVO'].tolist()
                tipo_logra1 = planilha['EnderecoTipo'].tolist()
                tipo_logra2 = planilha['EnderecoTipo_ARQUIVO'].tolist()
                titulo_logra1 = planilha['EnderecoTitulo'].tolist()
                titulo_logra2 = planilha['EnderecoTitulo_ARQUIVO'].tolist()
                num1 = planilha['Endereco_Numero'].tolist()
                num2 = planilha['Endereco_Numero_ARQUIVO'].tolist()
                compl1 = planilha['EnderecoComplemento'].tolist()
                compl2 = planilha['EnderecoComplemento_ARQUIVO'].tolist()
                bairros = planilha['Bairro'].tolist()
                bairros2 = planilha['Bairro_ARQUIVO'].tolist()
                cidade_uf1 = planilha['Estado - Cidade'].tolist()
                ceps1 = planilha['EnderecoCep'].tolist()
                ceps2 = planilha['EnderecoCep_ARQUIVO'].tolist()
                cidade2 = planilha['Cidade_ARQUIVO'].tolist()
                uf2 = planilha['Estado_ARQUIVO'].tolist()

                arq = arquivos_xlsx[x][5:-5]
                
                with open('path/count.txt', 'r+', encoding='UTF-8') as t:
                    count = t.read()
                if count == '0':
                    print()
                    print(f'Analisando planilha {arq}')
                    print()
                else:
                    print()
                    print(f'Continuando planilha {arq}')
                    print()

                with open('path/sheet.txt', 'w+', encoding='UTF-8') as t:
                    t.write(f'{x}')

                
                with open('path/count.txt', 'r+') as t:
                    ini = int(t.read())

                for i in range(ini, len(cpf1)):
                    while True:
                        try:
                            with open('path/count.txt', 'w+', encoding='UTF-8') as t:
                                t.write(f'{i}')

                            c = cpf1[i]
                            idade = idade1[i]
                            mae = maes[i]
                            nb = nb1[i]
                            
                            data_inde = str(data_inde1[i])
                            try:
                                dt = datetime.strptime(data_inde, '%Y-%m-%d %H:%M:%S')
                                data_inde = dt.strftime('%d/%m/%Y')
                            except ValueError:
                                pass

                            especie = especies[i]
                            motivo = motivos[i]

                            if pd.isna(nb):
                                nb = ''
                            if data_inde == 'nan':
                                data_inde = 'N/D'
                            if pd.isna(especie):
                                especie = ''
                            if pd.isna(motivo):
                                motivo = ''

                            orig_logra = logra1[i]
                            orig_logra2 = logra2[i]
                            orig_titulo = titulo_logra1[i]
                            orig_titulo2 = titulo_logra2[i]
                            orig_tipo_logra = tipo_logra1[i]
                            orig_tipo_logra2 = tipo_logra2[i]
                            orig_nasc = nasc1[i]
                            orig_dddcel = ddd_cel[i]
                            orig_dddcel1 = ddd_cel1[i]
                            orig_dddcel2 = ddd_cel2[i]
                            orig_cel = celular0[i]
                            orig_cel1 = celular1[i]
                            orig_cel2 = celular2[i]
                            ddd = ddds[i]
                            ddd1 = ddds1[i]
                            ddd2 = ddds2[i]
                            orig_tel = tel0[i]
                            orig_tel1 = tel1[i]
                            orig_tel2 = tel2[i]
                            orig_num = num1[i]
                            orig_num2 = num2[i]
                            orig_compl = compl1[i]
                            orig_compl2 = compl2[i]
                            orig_bair = bairros[i]
                            orig_bair2 = bairros2[i]
                            orig_cid_uf = cidade_uf1[i]
                            orig_cidade2 = cidade2[i]
                            orig_uf2 = uf2[i]
                            orig_cep = ceps1[i]
                            orig_cep2 = ceps2[i]

                            if pd.isna(orig_dddcel):
                                orig_dddcel = ''
                            if pd.isna(orig_dddcel1):
                                orig_dddcel1 = ''
                            if pd.isna(orig_dddcel2):
                                orig_dddcel2 = ''
                            if pd.isna(ddd):
                                ddd = ''
                            if pd.isna(ddd1):
                                ddd1 = ''
                            if pd.isna(ddd2):
                                ddd2 = ''
                            if pd.isna(orig_tel):
                                orig_tel = ''
                            if pd.isna(orig_tel1):
                                orig_tel1 = ''
                            if pd.isna(orig_tel2):
                                orig_tel2 = ''
                            if pd.isna(orig_cel):
                                orig_cel = ''
                            if pd.isna(orig_cel1):
                                orig_cel1 = ''
                            if pd.isna(orig_cel2):
                                orig_cel2 = ''

                            if pd.isna(orig_tipo_logra):
                                orig_tipo_logra = ''
                            if pd.isna(orig_titulo):
                                orig_titulo = ''
                            if pd.isna(orig_logra):
                                orig_logra = ''
                            
                            orig_logradouro = orig_logra
                            if orig_logradouro == '':
                                orig_logradouro = 'N/D'

                            if pd.isna(orig_tipo_logra2):
                                orig_tipo_logra2 = ''
                            if pd.isna(orig_titulo2):
                                orig_titulo2 = ''
                            if pd.isna(orig_logra2):
                                orig_logra2 = ''
                            
                            log_list2 = [orig_tipo_logra2, orig_titulo2, orig_logra2]
                            orig_logradouro2 = ''.join(log_list2).strip()
                            if orig_logradouro2 == '':
                                orig_logradouro2 = 'N/D'

                            if pd.isna(orig_uf2):
                                orig_uf2 = ''
                            if pd.isna(orig_cidade2):
                                orig_cidade2 = ''
                            
                            ciduf_list = [orig_uf2, orig_cidade2]
                            orig_cid_uf2 = ''.join(ciduf_list).strip()
                            if orig_cid_uf2 == '':
                                orig_cid_uf2 = 'N/D'

                            orig_celular = str(orig_dddcel) + str(orig_cel)
                            orig_celular1 = str(orig_dddcel1) + str(orig_cel1)
                            orig_celular2 = str(orig_dddcel2) + str(orig_cel2)

                            orig_tel = str(ddd) + str(orig_tel)
                            orig_tel1 = str(ddd1) + str(orig_tel1)
                            orig_tel2 = str(ddd2) + str(orig_tel2)
                            
                            c = list(c)

                            while len(c) < 11:
                                if len(c) == 11:
                                    print(c)
                                else:
                                    c = list(c)
                                    c.insert(0, '0')
                            c = ''.join(c)

                            cpf = c[:3]+'.'+c[3:6]+'.'+c[6:9]+'-'+c[-2:]

                            data = {
                                'javax.faces.partial.ajax': 'true',
                                'javax.faces.source': 'formAgendarConsultar:btnConsultarCpfAgendamento',
                                'javax.faces.partial.execute': 'formAgendarConsultar:btnConsultarCpfAgendamento formAgendarConsultar:cpfInput formAgendarConsultar:pnlSalarioMaternidadeWrapper',
                                'javax.faces.partial.render': 'formAgendarConsultar frmBotoes frmFormulariosServico',
                                'formAgendarConsultar:btnConsultarCpfAgendamento': 'formAgendarConsultar:btnConsultarCpfAgendamento',
                                'formAgendarConsultar': 'formAgendarConsultar',
                                'DTPINFRA_TOKEN': TOKEN,
                                'formAgendarConsultar:cpfInput': cpf,
                                'formAgendarConsultar:celularInput': '',
                                'formAgendarConsultar:fixoInput': '',
                                'formAgendarConsultar:telefoneSecundarioInput': '',
                                'formAgendarConsultar:emailInput': '',
                                'formAgendarConsultar:cepPrincipalInput': '',
                                'formAgendarConsultar:inputComboTipoLogradouro':'',
                                'formAgendarConsultar:logradouroPrincipalInput': '',
                                'formAgendarConsultar:numeroPrincipalInput': '',
                                'formAgendarConsultar:complementoPrincipalInput': '',
                                'formAgendarConsultar:bairroPrincipalInput': '',
                                'formAgendarConsultar:ufDrop2': '',
                                'formAgendarConsultar:checkEnderecoSecundario_input': 'on',
                                'formAgendarConsultar:paisEndSecundario': '76',
                                'formAgendarConsultar:cepSecundarioInput': '',
                                'formAgendarConsultar:inputComboTipoLogradouroSecundario': '',
                                'formAgendarConsultar:logradouroSecundarioInput': '',
                                'formAgendarConsultar:numeroSecundarioInput': '',
                                'formAgendarConsultar:complementoSecundarioInput': '',
                                'formAgendarConsultar:bairroSecundarioInput': '',
                                'formAgendarConsultar:ufDropSecundario2': '',
                                'javax.faces.ViewState': JSF
                            }

                            res = rq.post(url_p, headers=header, data=data)

                            soup2 = bs(res.content, 'lxml')

                            nome = soup2.find('input', {'id':'formAgendarConsultar:nomeInput'})
                            text_nome = nome.text

                            nasc = soup2.find('input', {'id':'formAgendarConsultar:nascimentoInput'})

                            # Informações de contato
                            celular = soup2.find('input', {'id':'formAgendarConsultar:celularInput'})
                            tel_prin = soup2.find('input', {'id':'formAgendarConsultar:fixoInput'})
                            tel_sec = soup2.find('input', {'id':'formAgendarConsultar:telefoneSecundarioInput'})
                            email = soup2.find('input', {'id':'formAgendarConsultar:emailInput'})

                            # Endereço principal
                            cep = soup2.find('input', {'id':'formAgendarConsultar:cepPrincipalInput'})
                            tip_log = soup2.find('select', {'id':'formAgendarConsultar:inputComboTipoLogradouro'})
                            logr = soup2.find('input', {'id':'formAgendarConsultar:logradouroPrincipalInput'})
                            numero =soup2.find('input', {'id':'formAgendarConsultar:numeroPrincipalInput'})
                            comp = soup2.find('input', {'id':'formAgendarConsultar:complementoPrincipalInput'})
                            bairro = soup2.find('input', {'id':'formAgendarConsultar:bairroPrincipalInput'})
                            estados = soup2.find('select', {'id':'formAgendarConsultar:ufDrop2'})
                            muni = soup2.find('input', {'id':'formAgendarConsultar:municipioEnderecoContato_input'})

                            # Endereço secundário
                            cep2 = soup2.find('input', {'id':'formAgendarConsultar:cepSecundarioInput'})
                            tip_log2 = soup2.find('select', {'id':'formAgendarConsultar:inputComboTipoLogradouroSecundario'})
                            logr2 = soup2.find('input', {'id':'formAgendarConsultar:logradouroSecundarioInput'})
                            numero2 = soup2.find('input', {'id':'formAgendarConsultar:numeroSecundarioInput'})
                            comp2 = soup2.find('input', {'id':'formAgendarConsultar:complementoSecundarioInput'})
                            bairro2 = soup2.find('input', {'id':'formAgendarConsultar:bairroSecundarioInput'})
                            estados2 = soup2.find('select', {'id':'formAgendarConsultar:ufDropSecundario2'})
                            muni2 = soup2.find('input', {'id':'formAgendarConsultar:municipioEnderecoSecundario_input'})
                            
                            try:
                                tipo_selected = tip_log.find('option', {'selected': 'selected'})
                                if tipo_selected.text == 'Selecione um tipo de logradouro':
                                    tipo_selected = None
                            except:
                                tipo_selected = None

                            try:
                                valor_logr = logr['value']
                            except:
                                valor_logr = None

                            if tipo_selected != None and valor_logr != None:
                                logradouro = f'{tipo_selected.text} {valor_logr}'
                            elif tipo_selected == None and valor_logr != None:
                                logradouro = valor_logr
                            elif tipo_selected != None and valor_logr == None:
                                logradouro = tipo_selected.text
                            elif tipo_selected == valor_logr == None:
                                logradouro = 'N/D'

                            estado_sel = estados.find('option', {'selected': 'selected'})

                            if estado_sel == None:
                                estado = 'N/D'
                            else:
                                estado = estado_sel.text[:2]

                            try:
                                muni = muni['value']
                            except:
                                muni = 'N/D'
                            try:
                                muni2 = muni2['value']
                            except:
                                muni2 = 'N/D'

                            estado_muni = f'{estado} - {muni}'
                            if 'N/D' in estado_muni:
                                estado_muni_list = estado_muni.split('N/D')
                                estado_muni = ''.join(estado_muni_list).strip()
                                if estado_muni == '-':
                                    estado_muni = 'N/D'

                            try:
                                tipo_selected2 = tip_log2.find('option', {'selected': 'selected'})
                                if tipo_selected2.text == 'Selecione um tipo de logradouro':
                                    tipo_selected2 = None
                            except:
                                tipo_selected2 = None

                            try:
                                valor_logr2 = logr2['value']
                            except:
                                valor_logr2 = None

                            try:
                                if tipo_selected2 != None and valor_logr2 != None:
                                    logradouro2 = f'{tipo_selected2.text} {valor_logr2}'
                                elif tipo_selected2 == None and valor_logr2 != None:
                                    logradouro2 = valor_logr2
                                elif tipo_selected2 != None and valor_logr2 == None:
                                    logradouro2 = tipo_selected2.text
                                elif tipo_selected2 == valor_logr2 == None:
                                    logradouro2 = 'N/D'
                            except:
                                logradouro2 = 'N/D'

                            try:
                                estado_sel2 = estados2.find('option', {'selected': 'selected'})
                                if estado_sel2 == None:
                                    estado2 = 'N/D'
                                else:
                                    estado2 = estado_sel2.text[:2]
                            except:
                                estado2 = 'N/D'

                            estado_muni2 = f'{estado2} - {muni2}'
                            if 'N/D' in estado_muni2:
                                estado_muni_list2 = estado_muni2.split('N/D')
                                estado_muni2 = ''.join(estado_muni_list2).strip()
                                if estado_muni2 == '-':
                                    estado_muni2 = 'N/D'

                            tags = [c, nome, nasc, idade, nb, data_inde, especie, motivo, mae, logradouro, numero, comp, bairro, estado_muni, cep, logradouro2, numero2, comp2, bairro2, estado_muni2, cep2, orig_logradouro, orig_num, orig_compl, orig_bair, orig_cid_uf, orig_cep, orig_logradouro2, orig_num2, orig_compl2, orig_bair2, orig_cid_uf2, orig_cep2, celular, tel_prin, tel_sec, orig_celular, orig_celular1, orig_celular2, orig_tel, orig_tel1, orig_tel2, email]
                            
                            valores = []

                            for valor in tags:
                                try:
                                    if type(valor) == str:
                                        valores.append(valor)
                                    else:
                                        valores.append(valor['value'])
                                except:
                                    valores.append('N/D')

                            nb = valores[4]
                            data_inde = valores[5]
                            especie = valores[6]
                            motivo = [7]
                            logradouro = valores[9]
                            bairro = valores[12]
                            estado_cidade = valores[13]
                            cep = valores[14]
                            endereco1 = [logradouro, bairro, estado_cidade, cep]

                            logradouro2 = valores[14]
                            bairro2 = valores[18]
                            estado_cidade2 = valores[19]
                            cep2 = valores[20]
                            endereco2 = [logradouro2, bairro2, estado_cidade2]

                            color = '00FFFFFF'
                            color2 = '00FFFFFF'
                            color3 = '00FFFFFF'

                            if len(cep) == 10:
                                n_cep = ''.join(cep.split('.'))
                                cep = ''.join(n_cep.split('-'))

                            mods = []
                            mods2 = []
                            mods3 = []

                            if 'N/D' in endereco1:

                                if endereco1.count('N/D') == 4 or endereco1.count('N/D') == 3:
                                    color = yellow
                                    # if valores[7] == 'N/D': 
                                    if pd.isna(orig_logradouro):
                                        orig_logradouro = 'N/D'
                                    valores[9] = orig_logradouro
                                    mods.append(9)
                                    # if valores[8] == 'N/D':
                                    if pd.isna(orig_num):
                                        orig_num = 'N/D'
                                    valores[10] = orig_num
                                    mods.append(10)
                                    # if valores[9] == 'N/D':
                                    if pd.isna(orig_compl):
                                        orig_compl = 'N/D'
                                    valores[11] = orig_compl
                                    mods.append(11)
                                    # if valores[10] == 'N/D':
                                    if pd.isna(orig_bair):
                                        orig_bair = 'N/D'
                                    valores[12] = orig_bair
                                    mods.append(12)
                                    # if valores[11] == 'N/D':
                                    if pd.isna(orig_cid_uf):
                                        orig_cid_uf = 'N/D'
                                    valores[13] = orig_cid_uf
                                    mods.append(13)
                                    # if valores[13] == 'N/D':
                                    if pd.isna(orig_cep):
                                        orig_cep = 'N/D'
                                    valores[14] = orig_cep
                                    mods.append(14)

                                if endereco1.count('N/D') == 2:
                                    color = yellow
                                    if valores[9] == 'N/D': 
                                        valores[9] = orig_logradouro
                                        mods.append(9)
                                    if valores[12] == 'N/D':
                                        valores[12] = orig_bair
                                        mods.append(12)
                                    if valores[13] == 'N/D':
                                        valores[13] = orig_cid_uf
                                        mods.append(13)
                                    if valores[14] == 'N/D':
                                        valores[14] = orig_cep
                                        mods.append(14)
                                
                                else:
                                    color = yellow
                                    if valores[9] == 'N/D': 
                                        if pd.isna(orig_logradouro):
                                            orig_logradouro = 'N/D'
                                        valores[9] = orig_logradouro
                                        mods.append(9)
                                    if valores[10] == 'N/D':
                                        if pd.isna(orig_num):
                                            orig_num = 'N/D'
                                        valores[10] = orig_num
                                        mods.append(10)
                                    if valores[11] == 'N/D':
                                        if pd.isna(orig_compl):
                                            orig_compl = 'N/D'
                                        valores[11] = orig_compl
                                        mods.append(11)
                                    if valores[12] == 'N/D':
                                        if pd.isna(orig_bair):
                                            orig_bair = 'N/D'
                                        valores[12] = orig_bair
                                        mods.append(12)
                                    if valores[13] == 'N/D':
                                        if pd.isna(orig_cid_uf):
                                            orig_cid_uf = 'N/D'
                                        valores[13] = orig_cid_uf
                                        mods.append(13)
                                    if valores[14] == 'N/D':
                                        if pd.isna(orig_cep):
                                            orig_cep = 'N/D'
                                        valores[14] = orig_cep
                                        mods.append(14)

                            if valores[1] == 'N/D':
                                raise
                                                    
                            try:
                                wb = load_workbook(f'results/{arq} RES.xlsx')
                                ws = wb.worksheets[0]
                                cells = ws[f'1:1']
                                if maior == []:
                                    for k in range(len(cells)):
                                        cell_size = len(cells[k].value)
                                        maior.append(cell_size)
                            except FileNotFoundError:
                                wb = Workbook()
                                ws = wb.active
                                ws.append(chaves)
                                cells = ws[f'1:1']
                                for k in range(len(cells)):
                                    cells[k].font = Font(bold=True)
                                    cells[k].alignment = Alignment(horizontal='center')
                                    cell_size = len(cells[k].value)
                                    maior.append(cell_size)
                                    ws.column_dimensions[cells[k].column_letter].width = cell_size+5

                            row = []

                            for f, valor in enumerate(valores):
                                try:
                                    if valor == '':
                                        valor = 'N/D'

                                    cell = WriteOnlyCell(ws, value=valor)

                                    if f in mods and valor != 'N/D':
                                        cell.fill = PatternFill(fgColor=color, fill_type='solid')

                                    elif f in mods2 and valor != 'N/D':
                                        cell.fill = PatternFill(fgColor=color2, fill_type='solid')
                                    
                                    if f in mods3:
                                        cell.fill = PatternFill(fgColor=color3, fill_type='solid')
                                except IndexError:
                                    continue
                
                                row.append(cell)

                            ws.append(row)
                            co += 1
                            cells = ws[f'{co}:{co}']
                            for k in range(len(cells)):
                                if cells[k].value != None:
                                    cell_size = len(cells[k].value)
                                    cells[k].alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
                                    if cell_size > maior[k]:
                                        if cell_size <= 85:
                                            maior[k] = cell_size
                                            ws.column_dimensions[cells[k].column_letter].width = maior[k]+10
                                        else:
                                            ws.column_dimensions[cells[k].column_letter].width = 85
                            wb.save(f'results/{arq} RES.xlsx')
                            
                            print(f'{i+1} - {valores[1]} - {valores[0]}')
                            break
                        except Exception as e:
                            print(e)
                            print(f'CPF: {cpf}')
                            print('Tentando novamente...')
                            sleep(1)
                            continue
                
                with open('path/count.txt', 'w+', encoding='UTF-8') as t:
                    t.write('0')
                print(f'Nova planilha {arq} RES completa!')
                break
            except Exception as e:
                print(e)
                continue

    with open('path/sheet.txt', 'w+', encoding='UTF-8') as t:
        t.write('0')

if __name__=='__main__':
    main()

